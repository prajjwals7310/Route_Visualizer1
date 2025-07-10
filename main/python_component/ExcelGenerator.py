import os
import re
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader
from dataclasses import dataclass
import json
import time
import psutil
import sys
REPROCESS ="--reprocess" in sys.argv

@dataclass
class AttributeSpec:
    name: str
    default: str
    spec: str

# Load key-value pairs from .properties file
def load_properties_file(path):
    config = {}
    if not os.path.exists(path):
        return config
    with open(path, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and '=' in line:
                key, val = line.split('=', 1)
                config[key.strip()] = val.strip()
    return config

# Resolve nested placeholders like {{key_{{innerKey}}}}
def resolve_all_nested(input_str, config):
    pattern = re.compile(r"\{\{([^{}]+)}}")
    seen = set()
    changed = True
    while changed and "{{" in input_str:
        changed = False
        def replacer(match):
            key = match.group(1).strip()
            if key in seen:
                return match.group(0)
            seen.add(key)
            val = config.get(key, f"{{{{{key}}}}}")
            nonlocal changed
            changed = True
            return val
        input_str = pattern.sub(replacer, input_str)
    return input_str

def resolve_dynamic_uri(uri, config):
    if "{{" not in uri:
        return uri
    return resolve_all_nested(resolve_all_nested(uri, config), config)

# Extract query params from resolved URI
def extract_query_params(uri, config):
    params = {}
    if "?" in uri:
        query = uri.split("?", 1)[1]
        for pair in query.split("&"):
            if "=" in pair:
                k, v = pair.split("=", 1)
                k = k.strip()
                v = v.strip()
                if v.startswith("{{") and v.endswith("}}"):
                    key = v[2:-2].strip()
                    v = config.get(key, "")
                params[k] = v
            else:
                k = pair.strip()
                params[k] = ""
    return params

# Format memory usage
def format_bytes(bytes_size):
    for unit in ["B", "KB", "MB", "GB"]:
        if bytes_size < 1024:
            return f"{bytes_size:.2f} {unit}"
        bytes_size /= 1024
    return f"{bytes_size:.2f} TB"

# Excel wrap style
def create_wrapped_style(cell):
    cell.alignment = Alignment(wrap_text=True)

# Write component file from Jinja2 template
def write_component_file(base_output, service, route_id, component, endpoint_name, direction, opts, spec_map):
    folder = base_output / "Component_Mapping" / service / route_id
    folder.mkdir(parents=True, exist_ok=True)

    # Construct proper filename like direct_AadharAuthentication_Producer.txt
    file_name = f"{component}_{endpoint_name}_{direction.lower()}.txt"
    file_path = folder / file_name

    with_spec = []
    without_spec = []

    for key, default in opts.items():
        spec = spec_map.get(key, "")
        line = f"| {key} | {default} | {spec} |"
        if spec:
            with_spec.append(line)
        else:
            without_spec.append(line)

            

    with open(file_path, "w") as f:
        f.write(f"Direction: {direction}\n")
        f.write(f"Component: {component}\n\n")
        f.write(f"| Attribute | Default | Specification |\n")
        f.write(f"|-----------|---------|----------------|\n")

        for line in with_spec + without_spec:
            f.write(f"{line}\n")


# Main execution
def main():
    total_start = time.time()
    mem_start = psutil.Process(os.getpid()).memory_info().rss

    print("\U0001F680 Report Generation Started")

    properties_path = Path("input/Path.text")
    config_map = load_properties_file(properties_path)

    xml_dir = config_map.get("input.xml.dir")
    json_path = config_map.get("output.json.path")
    output_excel_path = config_map.get("output.excel.path")
    processed_dir_path = config_map.get("processed.dir", "processed")

    if not (xml_dir and json_path and output_excel_path):
        print("❌ Required paths missing in Path.text")
        return

    output_base = Path(output_excel_path).parent
    excel_path = Path(output_excel_path)
    processed_dir = Path(processed_dir_path)

    output_base.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)

    with open(json_path) as f:
        component_json = json.load(f)

    workbook = Workbook()
    #workbook.active.title = "component_mapped"
    sheet_map = {}
    row_counter = {}
    serial_counter = {}

    headers = ["S.No", "Service", "Route", "Component", "Endpoint URI", "C/P", "Attribute", "Default", "Specification"]

    folders = [f for f in Path(xml_dir).iterdir() if f.is_dir()]
    if not folders:
        print(f"⚠️ No input folders found in {xml_dir}")
        return

    xml_count = 0
    route_count = 0
    endpoint_count = 0

    for folder in folders:
        pom_file = folder / "pom.xml"
        service = folder.name
        if pom_file.exists():
            tree = ET.parse(pom_file)
            root = tree.getroot()
            ns = {'m': root.tag.split('}')[0].strip('{')}
            artifact = root.find(".//m:artifactId", ns)
            if artifact is not None:
                service = artifact.text

        config_file = folder / "src/main/resources/application.properties"
        cfg = load_properties_file(config_file)

        for xml_file in Path(folder / "src/main/resources/routes").glob("*.xml"):
            xml_count += 1
            content = re.sub(r'<!--.*?-->', '', xml_file.read_text(), flags=re.DOTALL)
            root = ET.fromstring(content)

            for route in root.findall(".//{*}route"):
                route_count += 1
                route_id = route.attrib.get("id", f"route_{route_count}")

                for elem in route.iter():
                    tag = elem.tag.split("}")[-1]
                    if tag not in ("from", "to"):
                        continue

                    uri_raw = elem.attrib.get("uri", "")
                    if not uri_raw:
                        continue

                    uri_resolved = resolve_dynamic_uri(uri_raw, cfg)
                    if ":" not in uri_resolved:
                        continue

                    if ":queue:" in uri_resolved:
                        component = "activemq"
                    else:
                        component = uri_resolved.split(":", 1)[0].lower()

                    endpoint_name = uri_resolved.split(":", 1)[1].split("?")[0].replace("/", "").replace("=", "")
                    direction = "Consumer" if tag == "from" else "Producer"

                    comp_def = component_json.get(component, {})
                    dir_def = comp_def.get(direction, {})
                    opts = dir_def.get("endpoint_parameters", {})

                    spec_map = extract_query_params(uri_resolved, cfg)
                    write_component_file(output_base, service, route_id, component, endpoint_name, direction, opts, spec_map)
                    endpoint_count += 1

                    if component not in sheet_map:
                        sheet = workbook.create_sheet(component)
                        sheet_map[component] = sheet
                        sheet.append(headers)
                        for cell in sheet[1]:
                            cell.font = Font(bold=True)
                        row_counter[component] = 2
                        serial_counter[component] = 1
                    else:
                        sheet = sheet_map[component]

                    row = row_counter[component]
                    serial = serial_counter[component]

                    sheet.append([
                        serial, service, route_id, component,
                        f"<{tag} uri=\"{uri_resolved}\" />", "C" if tag == "from" else "P"
                    ])
                    create_wrapped_style(sheet.cell(row=row, column=5))
                    row += 1

                    sheet.append(["", "", "", "", "", "", "Endpoint Option"])
                    sheet.append(["", "", "", "", "", "", "Name", "Default", "Specification"])
                    row += 2

                    keys = sorted(set(opts.keys()).union(spec_map.keys()), key=lambda k: (0 if spec_map.get(k) else 1, k))
                    for key in keys:
                        sheet.append(["", "", "", "", "", "", key, opts.get(key, ""), spec_map.get(key, "")])
                        create_wrapped_style(sheet.cell(row=row + 1, column=9))
                        row += 1

                    row_counter[component] = row
                    serial_counter[component] += 1

                

        # destination = Path(processed_dir) / folder.name
        # shutil.move(str(folder), str(destination))
        # shutil.copytree(folder, destination, dirs_exist_ok=True)

        if not REPROCESS:
            destination = Path(processed_dir) / folder.name
            shutil.copytree(folder, destination, dirs_exist_ok=True)

    for sheet in sheet_map.values():
        for col in sheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = get_column_letter(column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

    if not sheet_map:
        sheet = workbook.active
        sheet.title = "NoData"
        sheet.append(["No data processed. Please check input files."])
    elif "Sheet" in workbook.sheetnames:
        std = workbook["Sheet"]
        if std.max_row == 1 and std.max_column == 1 and std["A1"].value is None:
            workbook.remove(std)

        workbook.save(excel_path)
    print(f"✅ Excel saved at: {excel_path}")

    total_end = time.time()
    mem_end = psutil.Process(os.getpid()).memory_info().rss
    print(f"\n📊 Summary:")
    print(f"📁 XML processed: {xml_count}")
    print(f"📦 Routes found: {route_count}")
    print(f"🧩 Endpoints written: {endpoint_count}")
    print(f"🗂 Reports at: {output_base/'Component_Mapping'}")
    print(f"📊 Excel at: {excel_path}")
    print(f"⏱ Time taken: {int((total_end - total_start) * 1000)} ms")
    print(f"🧠 Memory used: {format_bytes(mem_end - mem_start)}")

if __name__ == "__main__":
    main()
